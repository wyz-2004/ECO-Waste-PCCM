from __future__ import annotations

import argparse
import json
import os
import sys
from collections import defaultdict
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


SCRIPT_PATH = Path(__file__).resolve()
ROOT = Path(os.environ["ECOWASTE_PROJECT_ROOT"]).resolve() if os.environ.get("ECOWASTE_PROJECT_ROOT") else SCRIPT_PATH.parents[1]
if str(SCRIPT_PATH.parent) not in sys.path:
    sys.path.insert(0, str(SCRIPT_PATH.parent))

import ecowaste_core as base
import run_experiment as legacy


COUNTRY_XLSX = ROOT / "data" / "raw" / "What_a_Waste_3.0_COUNTRY_Dataset_&_Codebook.xlsx"
OUT_DIR = ROOT / "outputs" / "ecowaste_longterm"
TABLE_DIR = OUT_DIR / "tables"
PROTOCOL_DIR = OUT_DIR / "protocol"
SEEDS = [11, 23, 37, 53, 71, 89, 107, 131, 149, 173]
METHODS = ["Median", "KNN", "MICE", "MissForest", "SoftImpute", "Low-rank SVD", "PCCM"]
PROJECTED = [f"{method}+Proj" for method in METHODS]
EPS = 1e-9


def read_country_inputs() -> tuple[pd.DataFrame, pd.DataFrame]:
    framed = pd.read_excel(COUNTRY_XLSX, sheet_name="Country dataset", header=[0, 1])
    country = framed.copy()
    country.columns = [str(code).strip() for _, code in country.columns]
    country = country.dropna(how="all").reset_index(drop=True)
    country["city_code"] = country["iso3c"].astype(str)
    country["city_name"] = country["country_name"].astype(str)

    # Measurement-year and point-of-measurement fields are metadata, not completion targets.
    metadata = [
        col
        for col in country.columns
        if col.endswith("_year") or "point_of_measurement" in col.lower()
    ]
    country = country.drop(columns=metadata, errors="ignore")

    workbook = load_workbook(COUNTRY_XLSX, read_only=True, data_only=True)
    sheet = workbook["Codebook"]
    rows = sheet.iter_rows(values_only=True)
    header = [str(v).strip() if v is not None else "" for v in next(rows)]
    index = {name: i for i, name in enumerate(header) if name}
    wanted = {
        "country_name",
        "iso3c",
        "measurement",
        "point_of_measurement",
        "method_of_measurement",
        "additional_explanation_for_method_of_data_collection",
        "date_of_measurement",
        "source/reference",
        "source/reference_url",
        "notes",
    }
    compact: list[dict[str, object]] = []
    for values in rows:
        measurement = values[index["measurement"]] if index["measurement"] < len(values) else None
        iso3c = values[index["iso3c"]] if index["iso3c"] < len(values) else None
        if not measurement or not iso3c:
            continue
        compact.append({name: values[index[name]] if name in index and index[name] < len(values) else None for name in wanted})
    workbook.close()
    codebook = pd.DataFrame(compact).rename(
        columns={
            "iso3c": "city_code",
            "point_of_measurement": "point_of_measuremnet",
            "source/reference": "source",
            "source/reference_url": "weblink",
        }
    )
    explanation = codebook.pop("additional_explanation_for_method_of_data_collection").fillna("").astype(str)
    codebook["notes"] = codebook["notes"].fillna("").astype(str) + " " + explanation
    return country, codebook


def prepare_country_data() -> base.PreparedData:
    country, codebook = read_country_inputs()
    data = base.prepare_features(country, codebook)
    keep = [
        j
        for j, feature in enumerate(data.features)
        if base.group_id(feature)
        in {
            "generation",
            "composition",
            "collection",
            "treatment",
            "uncollected",
            "non_msw",
            "workers",
            "governance",
            "epr_drs",
            "separation",
        }
    ]
    return subset_features(data, keep)


def subset_features(data: base.PreparedData, keep: list[int]) -> base.PreparedData:
    features = [data.features[j] for j in keep]
    return base.PreparedData(
        city=data.city,
        codebook=data.codebook,
        features=features,
        percent_features=data.percent_features.intersection(features),
        binary_features=data.binary_features.intersection(features),
        raw_values=data.raw_values[features].copy(),
        model_values=data.model_values[:, keep],
        observed_mask=data.observed_mask[:, keep],
        feature_mean=data.feature_mean[keep],
        feature_scale=data.feature_scale[keep],
        static_reliability=data.static_reliability[:, keep],
        reliability=data.reliability[:, keep],
        evidence_features=data.evidence_features[:, keep, :],
        evidence_feature_names=data.evidence_feature_names,
        evidence_calibration_weights=data.evidence_calibration_weights,
        evidence_calibration_diagnostics=data.evidence_calibration_diagnostics,
        context=data.context,
    )


def make_masks(data: base.PreparedData, seed: int) -> dict[str, np.ndarray]:
    rng = np.random.default_rng(seed)
    eligible = data.observed_mask.copy()
    eligible[:, eligible.sum(axis=0) < 20] = False
    mcar = eligible & (rng.random(eligible.shape) < 0.20)

    region = data.context["region_id"].astype(str).to_numpy()
    income = data.context["income_id_2022"].astype(str).to_numpy()
    row_factor = pd.Series(region).map(pd.Series(region).value_counts(normalize=True)).to_numpy()
    income_factor = pd.Series(income).map(pd.Series(income).value_counts(normalize=True)).to_numpy()
    mar_prob = np.clip(0.08 + 0.55 * row_factor + 0.35 * income_factor, 0.08, 0.48)
    mar = eligible & (rng.random(eligible.shape) < mar_prob[:, None])

    values = np.nan_to_num(data.model_values, nan=-99.0)
    thresholds = np.nanpercentile(np.where(eligible, data.model_values, np.nan), 65, axis=0)
    high = values >= thresholds[None, :]
    mnar_prob = np.where(high, 0.34, 0.10)
    mnar = eligible & (rng.random(eligible.shape) < mnar_prob)

    block = base.make_block_mask(data, seed + 3)
    missingness = 1.0 - data.observed_mask.mean(axis=1)
    high_rows = missingness >= np.quantile(missingness, 0.75)
    high_missing = eligible & high_rows[:, None] & (rng.random(eligible.shape) < 0.35)
    return {
        "MCAR": mcar,
        "MAR-context": mar,
        "MNAR-high-value": mnar,
        "Subsystem-block": block,
        "High-missingness-country": high_missing,
    }


def method_prediction(data: base.PreparedData, train: np.ndarray, method: str, seed: int) -> tuple[np.ndarray, np.ndarray]:
    if method == "Median":
        pred = legacy.stat_imputer(data, train, "median")
        uncertainty = generic_uncertainty(data, train, pred)
    elif method == "KNN":
        pred = base.weighted_city_knn(data, train, k=10)
        uncertainty = generic_uncertainty(data, train, pred)
    elif method == "MICE":
        pred = legacy.mice_ridge_imputer(data, train, iterations=4, max_features=42)
        uncertainty = generic_uncertainty(data, train, pred)
    elif method == "MissForest":
        pred = legacy.missforest_imputer(data, train, seed=seed, iterations=2, trees=5)
        uncertainty = generic_uncertainty(data, train, pred)
    elif method == "SoftImpute":
        pred = legacy.softimpute(data, train, rank=18, shrink=0.16, epochs=32)
        uncertainty = generic_uncertainty(data, train, pred)
    elif method == "Low-rank SVD":
        pred = base.matrix_factorization(data, train, rank=14, epochs=34, seed=seed, use_evidence=False)
        uncertainty = generic_uncertainty(data, train, pred)
    elif method == "PCCM":
        pred, uncertainty = legacy.safe_pccm_predict(data, train, seed=seed, use_constraints=False)
    else:
        raise ValueError(method)
    fallback = legacy.stat_imputer(data, train, "median")
    return np.where(np.isfinite(pred), pred, fallback), np.maximum(np.nan_to_num(uncertainty, nan=1.0), 1e-4)


def generic_uncertainty(data: base.PreparedData, train: np.ndarray, pred: np.ndarray) -> np.ndarray:
    observed = np.where(train, data.model_values, np.nan)
    scale = np.nanstd(observed, axis=0)
    fallback = np.nanmedian(scale[np.isfinite(scale) & (scale > EPS)])
    scale = np.where(np.isfinite(scale) & (scale > EPS), scale, fallback if np.isfinite(fallback) else 1.0)
    missing = 1.0 - train.mean(axis=0)
    row_gap = 1.0 - train.mean(axis=1)
    return scale[None, :] * (0.55 + missing[None, :] + 0.25 * row_gap[:, None])


def feasibility(data: base.PreparedData, raw: np.ndarray, projected: np.ndarray) -> dict[str, float]:
    metrics = legacy.constraint_metrics(data, projected)
    metrics["projection_distance"] = float(np.mean(np.abs(projected - raw)))
    metrics["violation"] = float(
        np.nansum(
            [
                metrics.get("range_violation_magnitude", np.nan),
                metrics.get("composition_sum_abs_error", np.nan),
                metrics.get("treatment_sum_abs_error", np.nan),
            ]
        )
    )
    return metrics


def calibration(data: base.PreparedData, pred: np.ndarray, unc: np.ndarray, calib: np.ndarray, test: np.ndarray) -> dict[str, float]:
    return legacy.conformal_from_masks(data, pred, unc, calib, test, reliability_weighted=False)


def safe_evidence_calibration(
    data: base.PreparedData,
    pred: np.ndarray,
    unc: np.ndarray,
    calib: np.ndarray,
    test: np.ndarray,
    seed: int,
    use_evidence: bool,
) -> dict[str, float]:
    if not use_evidence:
        return calibration(data, pred, unc, calib, test)
    rng = np.random.default_rng(seed + 1881)
    entries = np.argwhere(calib)
    if len(entries) < 240:
        return calibration(data, pred, unc, calib, test)
    assignments = rng.integers(0, 3, size=len(entries))
    fold_masks = []
    for fold in range(3):
        tune = np.zeros_like(calib, dtype=bool)
        chosen = entries[assignments == fold]
        tune[chosen[:, 0], chosen[:, 1]] = True
        fold_masks.append((calib & ~tune, tune))
    gap = 1.0 - np.clip(np.nan_to_num(data.reliability, nan=0.55), 0.08, 1.0)
    centered = gap - float(np.mean(gap[calib]))
    best_beta = 0.0
    baseline_losses = []
    for fit, tune in fold_masks:
        result = legacy.conformal_from_masks(data, pred, unc, fit, tune, reliability_weighted=False)
        baseline_losses.append(float(result["ece"] + 0.002 * result["width_80"]))
    best_loss = float(np.mean(baseline_losses))
    for beta in [-1.0, -0.75, -0.50, -0.25, 0.25, 0.50, 0.75, 1.0, 1.25]:
        scaled = np.maximum(unc * np.exp(beta * centered), 1e-4)
        losses = []
        for fit, tune in fold_masks:
            result = legacy.conformal_from_masks(data, pred, scaled, fit, tune, reliability_weighted=False)
            losses.append(float(result["ece"] + 0.002 * result["width_80"]))
        loss = float(np.mean(losses))
        if loss < best_loss * 0.80 and sum(value < base for value, base in zip(losses, baseline_losses)) == 3:
            best_beta, best_loss = beta, loss
    result = calibration(data, pred, np.maximum(unc * np.exp(best_beta * centered), 1e-4), calib, test)
    result["evidence_scale_beta"] = best_beta
    return result


def aggregate(frame: pd.DataFrame, keys: list[str]) -> pd.DataFrame:
    numeric = [c for c in frame.columns if c not in set(keys + ["seed"]) and pd.api.types.is_numeric_dtype(frame[c])]
    grouped = frame.groupby(keys, dropna=False)
    mean = grouped[numeric].mean().reset_index()
    std = grouped[numeric].std(ddof=1).add_suffix("_std").reset_index()
    std = std.rename(columns={f"{key}_std": key for key in keys})
    return mean.merge(std, on=keys, how="left").merge(grouped.size().rename("seeds").reset_index(), on=keys, how="left")


def run_benchmark(data: base.PreparedData) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    rows: list[dict[str, object]] = []
    group_rows: list[dict[str, object]] = []
    mask_rows: list[dict[str, object]] = []
    groups = sorted({base.group_id(feature) for feature in data.features})
    for seed in SEEDS:
        for scenario, test in make_masks(data, seed).items():
            rng = np.random.default_rng(seed + 700)
            pool = data.observed_mask & ~test
            calib = pool & (rng.random(pool.shape) < 0.15)
            train = pool & ~calib
            for i, j in np.argwhere(test):
                mask_rows.append({"seed": seed, "scenario": scenario, "country_code": data.city.iloc[i]["iso3c"], "feature": data.features[j]})
            for method in METHODS:
                raw, unc = method_prediction(data, train, method, seed + 17)
                projected = legacy.exact_feasibility_projection(data, raw)
                for name, pred in [(method, raw), (f"{method}+Proj", projected)]:
                    result = base.evaluate(pred, data.model_values, test, None)
                    result["percent_mae_points"] = base.percent_mae_pp(data, pred, test)
                    result.update(feasibility(data, raw, pred))
                    result.update(calibration(data, pred, unc, calib, test))
                    result.update({"seed": seed, "scenario": scenario, "method": name})
                    rows.append(result)
                if scenario == "MCAR":
                    for group in groups:
                        group_mask = test & np.array([base.group_id(feature) == group for feature in data.features])[None, :]
                        result = base.evaluate(projected, data.model_values, group_mask, None)
                        result.update({"seed": seed, "method": f"{method}+Proj", "variable_group": group, "evaluated_cells": int(group_mask.sum())})
                        group_rows.append(result)
    return pd.DataFrame(rows), pd.DataFrame(group_rows), pd.DataFrame(mask_rows)


def run_country_tuned_upper_bound(data: base.PreparedData) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    modes = ["global", "group", "reliability", "group_reliability"]
    for seed in SEEDS:
        test = make_masks(data, seed)["MCAR"]
        rng = np.random.default_rng(seed + 1900)
        pool = data.observed_mask & ~test
        tune = pool & (rng.random(pool.shape) < 0.15)
        fit = pool & ~tune
        candidates = {}
        for mode in modes:
            pred, _ = legacy.pccm_predict(data, fit, seed=seed + 31, use_constraints=False, gate_mode=mode)
            candidates[mode] = float(np.mean(np.abs(pred[tune] - data.model_values[tune])))
        selected = min(candidates, key=candidates.get)
        raw, unc = legacy.pccm_predict(data, pool, seed=seed + 41, use_constraints=False, gate_mode=selected)
        pred = legacy.exact_feasibility_projection(data, raw)
        result = base.evaluate(pred, data.model_values, test, None)
        result.update(feasibility(data, raw, pred))
        result.update({"seed": seed, "selected_gate_mode": selected, "tuning_mae": candidates[selected]})
        rows.append(result)
    return pd.DataFrame(rows)


def run_country_evidence_ablation(data: base.PreparedData) -> pd.DataFrame:
    rows: list[dict[str, object]] = []
    variants = ["No evidence", "Shuffled evidence", "Heuristic evidence", "Learned evidence"]
    learned = data.reliability.copy()
    for seed in SEEDS:
        test = make_masks(data, seed)["MCAR"]
        rng = np.random.default_rng(seed + 2700)
        pool = data.observed_mask & ~test
        calib = pool & (rng.random(pool.shape) < 0.15)
        train = pool & ~calib
        for variant in variants:
            try:
                if variant == "No evidence":
                    data.reliability = np.full_like(learned, 0.55)
                    use_evidence = False
                elif variant == "Shuffled evidence":
                    data.reliability = rng.permutation(learned.ravel()).reshape(learned.shape)
                    use_evidence = True
                elif variant == "Heuristic evidence":
                    data.reliability = data.static_reliability.copy()
                    use_evidence = True
                else:
                    data.reliability = learned.copy()
                    use_evidence = True
                # Cross-scale evidence is allowed to adapt calibration only;
                # point completion and base uncertainty remain evidence-neutral.
                raw, unc = legacy.safe_pccm_predict(data, train, seed=seed + 47, use_evidence=False, use_constraints=False)
                pred = legacy.exact_feasibility_projection(data, raw)
                result = base.evaluate(pred, data.model_values, test, None)
                result.update(safe_evidence_calibration(data, pred, unc, calib, test, seed, use_evidence))
                result.update(feasibility(data, raw, pred))
                result.update({"seed": seed, "variant": variant})
                rows.append(result)
            finally:
                data.reliability = learned.copy()
    return pd.DataFrame(rows)


def paired_effects(raw: pd.DataFrame) -> pd.DataFrame:
    rng = np.random.default_rng(20260607)
    rows: list[dict[str, object]] = []
    projected = raw[raw.method.isin(PROJECTED)]
    for scenario in sorted(projected.scenario.unique()):
        pccm = projected[(projected.scenario == scenario) & (projected.method == "PCCM+Proj")].set_index("seed")
        for baseline in [m for m in PROJECTED if m != "PCCM+Proj"]:
            other = projected[(projected.scenario == scenario) & (projected.method == baseline)].set_index("seed")
            for metric in ["mae_norm", "rmse_norm", "projection_distance", "ece"]:
                paired = pd.concat([pccm[metric].rename("pccm"), other[metric].rename("baseline")], axis=1).dropna()
                diff = paired.pccm.to_numpy() - paired.baseline.to_numpy()
                boot = np.array([rng.choice(diff, len(diff), replace=True).mean() for _ in range(2000)])
                rows.append(
                    {
                        "scenario": scenario,
                        "comparison": f"PCCM+Proj vs {baseline}",
                        "metric": metric,
                        "paired_mean_difference": float(diff.mean()),
                        "ci95_low": float(np.quantile(boot, 0.025)),
                        "ci95_high": float(np.quantile(boot, 0.975)),
                        "wins": int((diff < 0).sum()),
                        "ties": int((np.abs(diff) < 1e-8).sum()),
                        "losses": int((diff > 0).sum()),
                    }
                )
    return pd.DataFrame(rows)


def group_summary(group_raw: pd.DataFrame) -> pd.DataFrame:
    stats = aggregate(group_raw, ["variable_group", "method"])
    rows = []
    for group, subset in stats.groupby("variable_group"):
        support = int(subset.evaluated_cells.max())
        eligible = support >= 30
        winner = subset.sort_values("mae_norm").iloc[0] if eligible else None
        pccm = subset[subset.method == "PCCM+Proj"].iloc[0]
        rank = int(subset.mae_norm.rank(method="min").loc[pccm.name]) if eligible else 0
        rows.append(
            {
                "variable_group": group,
                "support": support,
                "best_method": winner.method if winner is not None else "Not assigned",
                "best_mae": winner.mae_norm if winner is not None else np.nan,
                "pccm_mae": pccm.mae_norm,
                "pccm_rank": rank if eligible else np.nan,
                "support_status": "Supported" if eligible else "Insufficient support",
            }
        )
    return pd.DataFrame(rows).sort_values(["support_status", "support"], ascending=[False, False])


def write_schema(data: base.PreparedData) -> None:
    rows = []
    for j, feature in enumerate(data.features):
        rows.append(
            {
                "feature": feature,
                "variable_group": base.group_id(feature),
                "type": "binary" if feature in data.binary_features else "percentage" if feature in data.percent_features else "continuous",
                "observed_cells": int(data.observed_mask[:, j].sum()),
                "missing_rate": float(1.0 - data.observed_mask[:, j].mean()),
                "constraint": "simplex/closure" if base.group_id(feature) in {"composition", "treatment"} else "box [0,1]" if feature in data.percent_features else "soft audit/unconstrained",
            }
        )
    pd.DataFrame(rows).to_csv(PROTOCOL_DIR / "country_variable_constraint_map.csv", index=False, encoding="utf-8-sig")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--evidence-only", action="store_true")
    args = parser.parse_args()
    TABLE_DIR.mkdir(parents=True, exist_ok=True)
    PROTOCOL_DIR.mkdir(parents=True, exist_ok=True)
    data = prepare_country_data()
    write_schema(data)
    evidence_raw = run_country_evidence_ablation(data)
    evidence = aggregate(evidence_raw, ["variant"])
    evidence_raw.to_csv(TABLE_DIR / "country_evidence_ablation_by_seed.csv", index=False, encoding="utf-8-sig")
    evidence.to_csv(TABLE_DIR / "country_evidence_ablation_summary.csv", index=False, encoding="utf-8-sig")
    if args.evidence_only:
        print(evidence[["variant", "mae_norm", "coverage_80", "width_80", "ece"]].to_string(index=False))
        return
    raw, group_raw, masks = run_benchmark(data)
    summary = aggregate(raw, ["scenario", "method"])
    groups = group_summary(group_raw)
    tuned_raw = run_country_tuned_upper_bound(data)
    tuned = aggregate(tuned_raw, ["selected_gate_mode"])
    effects = paired_effects(raw)
    raw.to_csv(TABLE_DIR / "country_generalization_by_seed.csv", index=False, encoding="utf-8-sig")
    summary.to_csv(TABLE_DIR / "country_generalization_summary.csv", index=False, encoding="utf-8-sig")
    group_raw.to_csv(TABLE_DIR / "country_variable_group_by_seed.csv", index=False, encoding="utf-8-sig")
    groups.to_csv(TABLE_DIR / "country_variable_group_summary.csv", index=False, encoding="utf-8-sig")
    tuned_raw.to_csv(TABLE_DIR / "country_tuned_upper_bound_by_seed.csv", index=False, encoding="utf-8-sig")
    tuned.to_csv(TABLE_DIR / "country_tuned_upper_bound_summary.csv", index=False, encoding="utf-8-sig")
    effects.to_csv(TABLE_DIR / "country_paired_bootstrap_effects.csv", index=False, encoding="utf-8-sig")
    masks.to_csv(PROTOCOL_DIR / "country_shared_mask_assignments.csv", index=False, encoding="utf-8-sig")
    summary_info = {
        "countries": len(data.city),
        "modeled_features": len(data.features),
        "observed_rate": float(data.observed_mask.mean()),
        "evidence_records": len(data.codebook),
        "seeds": SEEDS,
        "methods": METHODS,
        "protocols": sorted(summary.scenario.unique().tolist()),
        "positioning": "same-program cross-scale generalization; not fully independent external validation",
        "zero_tuning": "All primary benchmark hyperparameters are inherited from the city experiment.",
    }
    (OUT_DIR / "country_generalization_summary.json").write_text(json.dumps(summary_info, indent=2), encoding="utf-8")
    print(json.dumps(summary_info, indent=2))


if __name__ == "__main__":
    main()
